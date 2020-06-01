import discord
import asyncio
import openpyxl
import json
from discord import Member
from discord.ext import commands
from bs4 import BeautifulSoup
from urllib.request import urlopen
from datetime import datetime
from modules import getdata
client = discord.Client()

## Setting
member = ['narinn-star', 'min-0', 'taehyundev','Good-jjun','lee-hanju']


## Event
@client.event
async def on_ready():
    print(client.user.id)
    print("ready")
    game = discord.Game("식사")
    await client.change_presence(status=discord.Status.online, activity=game)

@client.event
async def on_message(message):
    if message.content.startswith("#help"):
        await message.channel.send("< Menual >")
        embed = discord.Embed(
        title = '명령어',
        colour = discord.Colour.blue()
        )
        embed.add_field(value = '@github <닉네임>', name = 'github에 커밋을 얼마나 올렸는지를 체크할 수 있다.\n',inline = False)
        embed.add_field(value='!기억초기화 or !기억 초기화', name='학습시킨 데이터를 모두 초기화 시킬 수 있다.\n', inline=False)
        embed.add_field(value='!학습 [질문] [답변]', name='질문에 대응하여 답변을 하게끔 학습을 시킬 수 있다.\n', inline=False)
        embed.add_field(value='+ [질문]', name='narinBot에게 질문을 할 수 있다.\n', inline=False)
        await message.channel.send(embed=embed)
            
    ### github   
    if message.content.startswith("@github"):
        try:
            name = message.content.split(" ")
            #tier System
            if name[1] == '-rank':
                if len(name) ==3:
                    score = getdata.getContribution_one(name[2])
                    tier = getdata.getRank_one(score)
                    await message.channel.send(name[2] + "님의 티어는 "+tier+"입니다.!!")
                elif len(name) ==2:
                    info = getdata.getContribution(member)
                    tierinfo = getdata.getRank(info)
                    name_keys = tierinfo.keys()
                    for name in name_keys:
                        await message.channel.send(name + "님의 티어는 "+tierinfo[name]+"입니다.!!")
            elif name[1] == '-all':
                for i in range(len(member)):
                    datacount =getdata.getCommitData(member[i])
                    print(datacount)
                    if datacount == 0:
                        await message.channel.send(member[i]+"님, 커밋을 하나도 안했네요??..?") 
                    else:
                        await message.channel.send(member[i]+"님,커밋을 "+str(datacount)+" 만큼 올리셨군요!")
            else:
                datacount =getdata.getCommitData(name[1])
                if datacount == 0:
                    await message.channel.send(name[1]+"님, 커밋을 하나도 안했네요??..?") 
                else:
                    await message.channel.send(name[1]+"님,커밋을 "+str(datacount)+" 만큼 올리셨군요!")
        except:
            await message.channel.send("값을 잘못입력하셨군요..!")


    ## 데이터 학습
    if message.content.startswith("!기억 초기화") or message.content.startswith("!기억초기화"):
        file = openpyxl.load_workbook("memory.xlsx")
        sheet = file.active
        for i in range(1, 251):
            sheet["A"+str(i)].value = "-"
        await message.channel.send("기억초기화 완료")
        file.save("memory.xlsx")

    if message.content.startswith("!학습"):
        file = openpyxl.load_workbook('memory.xlsx')
        sheet = file.active
        learn = message.content.split(":")

        learn[0] = learn[0].replace("!학습 ", "")
        for i in range(1, 201):
            if sheet["A"+str(i)].value == "-":
              #  await client.send_message(message.channel, "기억되었습니당")
                sheet["A" + str(i)].value = learn[0]
                sheet["B" + str(i)].value = learn[1]
                break
        file.save("memory.xlsx")
    if message.content.startswith("!리스트"):
        await message.channel.send("<현재 학습리스트>")
        file = openpyxl.load_workbook('memory.xlsx')
        sheet = file.active
        data = ""

        for i in range(1, 201):
            if sheet["A" + str(i)].value == "-" and i == 1:
                await message.channel.send(message.channel, "데이터 없음")
            if sheet["A" + str(i)].value == "-":
                break
            data = data + "질문 : "+sheet["A" + str(i)].value + ", 답변 : "+ sheet["B" + str(i)].value + "\n"
        await message.channel.send(data)  
        print("Good ! ")
    if message.content.startswith("+ "):
        file = openpyxl.load_workbook('memory.xlsx')
        sheet = file.active
        memory = message.content
        memory = memory.replace("+ ", "")
        search = False
        for i in range(1, 201):
            if sheet["A" + str(i)].value == memory:
                await message.channel.send(sheet["B" + str(i)].value)
                search = True
                break
        
        if search == False:
            await message.channel.send("무슨 뜻인지 모르겠어요..")
        else:
            search = False
    if message.content.startswith('A'):
        if(message.content.startswith('B')):
            print('a')
        print('aaaa')


if __name__ == "__main__":
    with open('data/token.json', 'r') as f:
        json_data = json.load(f)
    token = json_data["value"]
    client.run(token)
